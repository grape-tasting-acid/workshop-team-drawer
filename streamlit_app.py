import csv
import io
import random
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import streamlit as st
from openpyxl import Workbook
st.set_page_config(page_title="조 추첨기", layout="wide")

DATA_DIR = Path(__file__).parent / "data"
OUTPUT_DIR = Path(__file__).parent / "output"

GROUP_COLORS = {"ob": "#1f77b4", "yb": "#2ca02c", "girls": "#d62728"}
DISPLAY_LABELS = {"leader": "CAPTAIN", "ob": "VETERANS", "yb": "ROOKIES", "girls": "GIRLS"}

# 재밌는 멘트
QUIPS = [
    "이 분이 팀의 분위기 메이커?!",
    "오늘의 럭키가이(걸)!",
    "캡틴이 살짝 긴장하는 표정입니다...",
    "운명은 이미 정해져 있었다...",
    "오오오~~ 예상 밖의 조합!",
    "팀워크 시너지가 느껴진다!",
    "커피 쿠폰 걸고 달려봅시다 ☕️",
    "박수 갈까요? 👏",
    "환호 부탁드립니다! 🙌",
]


def toast(msg: str) -> None:
    try:
        st.toast(msg)
    except Exception:
        st.caption(msg)


@dataclass
class Member:
    name: str
    group: str  # one of {"leader", "ob", "yb", "girls"}
    gender: str  # one of {"M", "F"}


@dataclass
class Team:
    index: int
    leader: Member
    members: List[Member] = field(default_factory=list)

    def add_member(self, member: Member) -> None:
        self.members.append(member)

    def all_people(self) -> List[Member]:
        return [self.leader] + self.members


@dataclass
class Room:
    index: int
    members: List[Member] = field(default_factory=list)


def read_leaders_csv_from_bytes(data: bytes) -> List[Member]:
    leaders: List[Member] = []
    f = io.StringIO(data.decode("utf-8"))
    reader = csv.DictReader(f)
    for row in reader:
        name = (row.get("name") or row.get("Name") or "").strip()
        gender = (row.get("gender") or row.get("Gender") or "").strip().upper()
        if not name:
            continue
        if gender not in {"M", "F"}:
            raise ValueError(f"캡틴 성별은 M/F 로 표기해야 합니다: {name} -> {gender}")
        leaders.append(Member(name=name, group="leader", gender=gender))
    return leaders


def read_names_csv_from_bytes(data: bytes, group: str, gender: str) -> List[Member]:
    members: List[Member] = []
    f = io.StringIO(data.decode("utf-8"))
    reader = csv.DictReader(f)
    for row in reader:
        name = (row.get("name") or row.get("Name") or "").strip()
        if not name:
            continue
        members.append(Member(name=name, group=group, gender=gender))
    return members


def read_csv_from_disk(path: Path) -> bytes:
    if not path.exists():
        return b""
    return path.read_bytes()


def compute_balanced_targets(
    leaders: List[Member],
    ob_count: int,
    yb_count: int,
    girls_count: int,
    rng: random.Random,
) -> Dict[int, Dict[str, int]]:
    """팀 총 인원(리더 제외)을 최대한 균등하게 맞추되,
    - 여성을 먼저 균등화(여리더 포함 여성 총합 균등) 후 `girls` 목표를 정한다.
    - 남성 그룹(`yb` → `ob`)을 팀별로 최대 균등(base+1 분배)하게 배정한다.
    - 남성 그룹 배정 시 여리더 팀을 우선 순서로 라운드로빈하되, 팀 용량(capacity)을 준수한다.
    - 목표: 각 그룹별 팀 간 격차를 최소화(현실적 제약 하에서 차이 2 이내 수렴)
    """
    num_teams = len(leaders)
    total = ob_count + yb_count + girls_count

    base = total // num_teams
    remainder = total % num_teams

    # 여분 1명을 받는 팀을 무작위로 선정
    extra_indices = rng.sample(range(num_teams), remainder) if remainder > 0 else []
    remaining_capacity = [base + (1 if i in extra_indices else 0) for i in range(num_teams)]

    targets: Dict[int, Dict[str, int]] = {i: {"ob": 0, "yb": 0, "girls": 0} for i in range(num_teams)}

    male_leader_idx = [i for i, ld in enumerate(leaders) if ld.gender == "M"]
    female_leader_idx = [i for i, ld in enumerate(leaders) if ld.gender == "F"]

    # 우선순위 리스트도 무작위 셔플
    rng.shuffle(male_leader_idx)
    rng.shuffle(female_leader_idx)
    others_from_female = [i for i in range(num_teams) if i not in female_leader_idx]
    others_from_male = [i for i in range(num_teams) if i not in male_leader_idx]
    rng.shuffle(others_from_female)
    rng.shuffle(others_from_male)

    order_for_male_groups = female_leader_idx + others_from_female  # ob/yb는 여리더 우선

    # 1) 팀별 최종 여성 수 목표치 산정(여리더 포함)
    total_females = girls_count + len(female_leader_idx)
    base_f = total_females // num_teams
    rem_f = total_females % num_teams
    # 여분 여성 1명을 받을 팀: 남리더 팀을 우선 선정 후 남으면 나머지 팀 무작위
    male_first_order = male_leader_idx + [i for i in range(num_teams) if i not in male_leader_idx]
    rng.shuffle(male_first_order)
    extra_female_indices = set(male_first_order[:rem_f]) if rem_f > 0 else set()
    desired_female_totals = [base_f + (1 if i in extra_female_indices else 0) for i in range(num_teams)]

    # 2) 각 팀 Girls 목표치 산정 = 최종 여성 목표치 - (리더가 여자면 1, 아니면 0)
    girls_targets = [max(0, desired_female_totals[i] - (1 if i in female_leader_idx else 0)) for i in range(num_teams)]

    # 3) 좌석(capacity) 초과/부족 보정 및 전체 girls_count에 맞춰 조정
    # 우선 capacity에 맞춰 캡
    girls_targets = [min(girls_targets[i], remaining_capacity[i]) for i in range(num_teams)]
    current_sum = sum(girls_targets)
    # 부족분을 남은 좌석이 있는 팀에 배분(남리더 우선)
    if current_sum < girls_count:
        need = girls_count - current_sum
        fill_order = male_leader_idx + [i for i in range(num_teams) if i not in male_leader_idx]
        # 라운드로빈 배분
        ptr = 0
        while need > 0 and any(remaining_capacity[i] - girls_targets[i] > 0 for i in range(num_teams)):
            team_i = fill_order[ptr % len(fill_order)] if fill_order else ptr % num_teams
            if remaining_capacity[team_i] - girls_targets[team_i] > 0:
                girls_targets[team_i] += 1
                need -= 1
            ptr += 1
    elif current_sum > girls_count:
        # 초과분을 줄임(여리더 팀을 우선 줄여 총 여성 균등을 유지)
        over = current_sum - girls_count
        reduce_order = female_leader_idx + [i for i in range(num_teams) if i not in female_leader_idx]
        ptr = 0
        while over > 0 and any(g > 0 for g in girls_targets):
            team_i = reduce_order[ptr % len(reduce_order)] if reduce_order else ptr % num_teams
            if girls_targets[team_i] > 0:
                girls_targets[team_i] -= 1
                over -= 1
            ptr += 1

    # girls 목표치 확정 적용
    for i in range(num_teams):
        targets[i]["girls"] = girls_targets[i]
        remaining_capacity[i] -= girls_targets[i]

    def allocate_group_even(count: int, preferred_order: List[int], key: str) -> None:
        if count <= 0:
            return
        num = len(remaining_capacity)
        if num == 0:
            return
        base = count // num
        # 1) 팀별 base 만큼 배정(용량 한도 고려)
        allocated = 0
        for i in range(num):
            give = min(base, max(0, remaining_capacity[i]))
            targets[i][key] += give
            remaining_capacity[i] -= give
            allocated += give
        left = count - allocated
        if left <= 0:
            return
        # 2) 여분을 선호 순서(여리더 우선)로 라운드로빈 배정(용량 남은 팀만)
        order = preferred_order + [i for i in range(num) if i not in preferred_order]
        ptr = rng.randrange(len(order)) if order else 0
        while left > 0 and any(remaining_capacity[i] > 0 for i in range(num)):
            team_i = order[ptr % len(order)] if order else (ptr % num)
            if remaining_capacity[team_i] > 0:
                targets[team_i][key] += 1
                remaining_capacity[team_i] -= 1
                left -= 1
            ptr += 1

    # 4) 남은 좌석에 남성 그룹을 균등 배정: YB → OB (여리더 우선)
    allocate_group_even(yb_count, order_for_male_groups, "yb")
    allocate_group_even(ob_count, order_for_male_groups, "ob")

    return targets


def assign_members_to_teams(
    leaders: List[Member],
    ob_list: List[Member],
    yb_list: List[Member],
    girls_list: List[Member],
    seed: Optional[int] = None,
) -> List[Team]:
    if seed is None:
        seed = int(time.time() * 1000) % (2**32 - 1)
    rng = random.Random(seed)

    num_teams = len(leaders)
    if num_teams != 8:
        raise ValueError("캡틴 수는 반드시 8명이어야 합니다.")
    male_count = sum(1 for m in leaders if m.gender == "M")
    female_count = sum(1 for m in leaders if m.gender == "F")
    if male_count != 4 or female_count != 4:
        raise ValueError("캡틴 성별은 남 4, 여 4 이어야 합니다.")

    targets = compute_balanced_targets(
        leaders, ob_count=len(ob_list), yb_count=len(yb_list), girls_count=len(girls_list), rng=rng
    )

    rng.shuffle(ob_list)
    rng.shuffle(yb_list)
    rng.shuffle(girls_list)

    teams = [Team(index=i, leader=leaders[i]) for i in range(num_teams)]

    def pop_many(src: List[Member], count: int) -> List[Member]:
        if count <= 0:
            return []
        if count > len(src):
            raise ValueError("분배 대상 인원이 부족합니다. 입력 CSV를 확인하세요.")
        result = src[:count]
        del src[:count]
        return result

    for i in range(num_teams):
        teams[i].members.extend(pop_many(ob_list, targets[i]["ob"]))
    for i in range(num_teams):
        teams[i].members.extend(pop_many(yb_list, targets[i]["yb"]))
    for i in range(num_teams):
        teams[i].members.extend(pop_many(girls_list, targets[i]["girls"]))

    for team in teams:
        rng.shuffle(team.members)

    # 특별 룰: 특정 두 사람은 같은 팀이 되지 않도록 스왑으로 조정
    enforce_exclusion_pairs(teams, pairs=[("노시현", "배연경")])
    enforce_exclusion_pairs(teams, pairs=[("노시현", "이진원")])

    # 포함 룰: seed가 3의 배수면 이진원-배연경은 같은 팀이 되도록 조정
    try:
        if seed is not None and seed % 3 == 0:
            enforce_inclusion_pair(teams, pair=("이진원", "배연경"))
    except Exception:
        pass

    # 포함 룰로 인해 분포가 흐트러졌을 수 있으므로, 목표치(targets)에 맞춰 재균형
    try:
        desired_girls_by_team: List[int] = [targets[i]["girls"] for i in range(num_teams)]
        desired_yb_by_team: List[int] = [targets[i]["yb"] for i in range(num_teams)]
        desired_ob_by_team: List[int] = [targets[i]["ob"] for i in range(num_teams)]
        exclude_names = {"이진원", "배연경", "노시현"}

        # 포함/제외로 인해 이동 불가한 YB(락 인원)의 최소 보장치 반영 → 목표치 보정
        try:
            locked_yb_by_team: List[int] = []
            for t in teams:
                locked_yb = sum(1 for m in t.members if m.group == "yb" and m.name in exclude_names)
                locked_yb_by_team.append(locked_yb)
            yb_total_before = sum(desired_yb_by_team)
            ob_total_before = sum(desired_ob_by_team)
            # 팀별 하한 충족(락 수만큼 yb 목표 상향, 동일 수만큼 ob 목표 하향)
            for i in range(num_teams):
                if locked_yb_by_team[i] > desired_yb_by_team[i]:
                    delta = locked_yb_by_team[i] - desired_yb_by_team[i]
                    desired_yb_by_team[i] += delta
                    take = min(delta, desired_ob_by_team[i])
                    desired_ob_by_team[i] -= take
            # 총합 보정: yb 총합이 늘었으면 다른 팀에서 yb를 1씩 줄이고 ob를 1씩 늘림(락 미포함 팀 우선)
            yb_excess = sum(desired_yb_by_team) - yb_total_before
            if yb_excess > 0:
                order = list(range(num_teams))
                # 락만큼의 하한이 없는 팀을 우선 감소 대상으로 선정
                order.sort(key=lambda i: (locked_yb_by_team[i] >= desired_yb_by_team[i], desired_yb_by_team[i]))
                ptr = 0
                while yb_excess > 0 and ptr < 256:
                    i = order[ptr % len(order)]
                    # 락 하한보다 큰 경우에만 감소 가능
                    if desired_yb_by_team[i] > locked_yb_by_team[i]:
                        desired_yb_by_team[i] -= 1
                        desired_ob_by_team[i] += 1
                        yb_excess -= 1
                    ptr += 1
            # 안전: 음수 방지 및 총합 유지
            for i in range(num_teams):
                if desired_ob_by_team[i] < 0:
                    # 부족분은 임의 팀에서 줄여 채움
                    need = -desired_ob_by_team[i]
                    desired_ob_by_team[i] = 0
                    # 다른 팀의 ob에서 차감하고 yb를 증가
                    j_ptr = 0
                    order2 = list(range(num_teams))
                    while need > 0 and j_ptr < 256:
                        j = order2[j_ptr % len(order2)]
                        if j != i and desired_ob_by_team[j] > 0:
                            desired_ob_by_team[j] -= 1
                            desired_yb_by_team[j] += 1
                            need -= 1
                        j_ptr += 1
        except Exception:
            pass

        def current_girls_counts() -> List[int]:
            return [sum(1 for m in t.members if m.group == "girls") for t in teams]

        cur = current_girls_counts()
        # 1) Girls: 반복적으로 도너(초과) → 리시버(부족) 간 스왑 시도
        safety = 0
        while True:
            safety += 1
            if safety > 64:
                break
            donors = [i for i, c in enumerate(cur) if c > desired_girls_by_team[i]]
            receivers = [i for i, c in enumerate(cur) if c < desired_girls_by_team[i]]
            if not donors or not receivers:
                break
            progressed = False
            for di in donors:
                # donor에서 교환할 girls 후보(제외 인원 제외)
                dg_idx = next((idx for idx, mem in enumerate(teams[di].members)
                                if mem.group == "girls" and mem.name not in exclude_names), None)
                if dg_idx is None:
                    continue
                for ri in receivers:
                    # receiver에서 교환할 남성 그룹 후보
                    rm_idx = next((idx for idx, mem in enumerate(teams[ri].members)
                                    if mem.group in {"ob", "yb"} and mem.name not in exclude_names), None)
                    if rm_idx is None:
                        continue
                    # 스왑 실행
                    teams[di].members[dg_idx], teams[ri].members[rm_idx] = teams[ri].members[rm_idx], teams[di].members[dg_idx]
                    # 카운트 업데이트
                    cur[di] -= 1
                    cur[ri] += 1
                    progressed = True
                    break
                if progressed:
                    break
            if not progressed:
                break

        # 2) YB: 목표치에 맞춰 균형(도너의 yb 한 명 ↔ 리시버의 ob 한 명 교환)
        def current_counts_for(group: str) -> List[int]:
            return [sum(1 for m in t.members if m.group == group) for t in teams]

        cur_yb = current_counts_for("yb")
        safety = 0
        while True:
            safety += 1
            if safety > 64:
                break
            donors = [i for i, c in enumerate(cur_yb) if c > desired_yb_by_team[i]]
            receivers = [i for i, c in enumerate(cur_yb) if c < desired_yb_by_team[i]]
            if not donors or not receivers:
                break
            progressed = False
            for di in donors:
                dy_idx = next((idx for idx, mem in enumerate(teams[di].members)
                               if mem.group == "yb" and mem.name not in exclude_names), None)
                if dy_idx is None:
                    continue
                for ri in receivers:
                    ro_idx = next((idx for idx, mem in enumerate(teams[ri].members)
                                   if mem.group == "ob" and mem.name not in exclude_names), None)
                    if ro_idx is None:
                        continue
                    teams[di].members[dy_idx], teams[ri].members[ro_idx] = teams[ri].members[ro_idx], teams[di].members[dy_idx]
                    cur_yb[di] -= 1
                    cur_yb[ri] += 1
                    progressed = True
                    break
                if progressed:
                    break
            if not progressed:
                break

        # 3) OB: 목표치에 맞춰 균형(도너의 ob 한 명 ↔ 리시버의 yb 한 명 교환)
        cur_ob = current_counts_for("ob")
        safety = 0
        while True:
            safety += 1
            if safety > 64:
                break
            donors = [i for i, c in enumerate(cur_ob) if c > desired_ob_by_team[i]]
            receivers = [i for i, c in enumerate(cur_ob) if c < desired_ob_by_team[i]]
            if not donors or not receivers:
                break
            progressed = False
            for di in donors:
                do_idx = next((idx for idx, mem in enumerate(teams[di].members)
                               if mem.group == "ob" and mem.name not in exclude_names), None)
                if do_idx is None:
                    continue
                for ri in receivers:
                    ry_idx = next((idx for idx, mem in enumerate(teams[ri].members)
                                   if mem.group == "yb" and mem.name not in exclude_names), None)
                    if ry_idx is None:
                        continue
                    teams[di].members[do_idx], teams[ri].members[ry_idx] = teams[ri].members[ry_idx], teams[di].members[do_idx]
                    cur_ob[di] -= 1
                    cur_ob[ri] += 1
                    progressed = True
                    break
                if progressed:
                    break
            if not progressed:
                break

        # 4) 마지막 안전장치: 목표치로 OB/YB를 재구성하여 정확히 맞춤(여성 수/락 인원은 유지)
        try:
            def rebuild_ob_yb_exact(locked_names: set[str]) -> None:
                ob_pool: List[Member] = []
                yb_pool: List[Member] = []
                # 팀별로 잠금 인원과 girls는 유지, 잠금 아닌 ob/yb는 풀로 회수
                for t in teams:
                    keep: List[Member] = []
                    for m in t.members:
                        if m.group in {"ob","yb"} and m.name not in locked_names:
                            if m.group == "ob":
                                ob_pool.append(m)
                            else:
                                yb_pool.append(m)
                        else:
                            keep.append(m)
                    t.members = keep
                rng.shuffle(ob_pool)
                rng.shuffle(yb_pool)
                # 팀별로 잠금된 ob/yb를 고려한 필요량 계산 후 채움
                for i, t in enumerate(teams):
                    locked_yb_here = sum(1 for m in t.members if m.group == "yb")
                    locked_ob_here = sum(1 for m in t.members if m.group == "ob")
                    need_yb = max(0, desired_yb_by_team[i] - locked_yb_here)
                    need_ob = max(0, desired_ob_by_team[i] - locked_ob_here)
                    take = min(need_yb, len(yb_pool))
                    t.members.extend(yb_pool[:take])
                    del yb_pool[:take]
                    take = min(need_ob, len(ob_pool))
                    t.members.extend(ob_pool[:take])
                    del ob_pool[:take]
                # 남은 경우가 드물지만, 좌석/목표에 맞춰 라운드로빈
                def fill_leftovers(pool: List[Member]) -> None:
                    if not pool:
                        return
                    order = list(range(num_teams))
                    rng.shuffle(order)
                    ptr = 0
                    while pool and ptr < 1024:
                        i = order[ptr % len(order)]
                        current_ob = sum(1 for m in teams[i].members if m.group == "ob")
                        current_yb = sum(1 for m in teams[i].members if m.group == "yb")
                        target_total = desired_yb_by_team[i] + desired_ob_by_team[i]
                        current_total = current_ob + current_yb
                        if current_total < target_total:
                            teams[i].members.append(pool.pop())
                        ptr += 1
                fill_leftovers(yb_pool)
                fill_leftovers(ob_pool)

            # 4-1) 1차 재구성(락: 포함/제외 예외 인원)
            rebuild_ob_yb_exact(exclude_names)
        except Exception:
            pass

        # 5) 재구성으로 인해 특수 포함/제외 룰 위반이 생길 수 있으므로 재적용(같은 그룹 우선 스왑)
        try:
            enforce_exclusion_pairs(teams, pairs=[("노시현", "배연경")])
            enforce_exclusion_pairs(teams, pairs=[("노시현", "이진원")])
            if seed is not None and seed % 3 == 0:
                enforce_inclusion_pair(teams, pair=("이진원", "배연경"))
        except Exception:
            pass

        # 6) 포함 룰로 인해 girls 분포가 살짝 흔들릴 수 있어 girls를 한 번 더 미세 보정
        try:
            desired_girls_by_team = [targets[i]["girls"] for i in range(num_teams)]
            cur = current_girls_counts()
            donors = [i for i, c in enumerate(cur) if c > desired_girls_by_team[i]]
            receivers = [i for i, c in enumerate(cur) if c < desired_girls_by_team[i]]
            for di in donors:
                dg_idx = next((idx for idx, mem in enumerate(teams[di].members)
                                if mem.group == "girls" and mem.name not in exclude_names), None)
                if dg_idx is None:
                    continue
                for ri in receivers:
                    rm_idx = next((idx for idx, mem in enumerate(teams[ri].members)
                                   if mem.group in {"ob","yb"} and mem.name not in exclude_names), None)
                    if rm_idx is None:
                        continue
                    teams[di].members[dg_idx], teams[ri].members[rm_idx] = teams[ri].members[rm_idx], teams[di].members[dg_idx]
                    break
        except Exception:
            pass

        # 7) 포함/제외 재적용 이후, 락을 포함해 정확히 목표치로 재구성(최종 보정)
        try:
            locked_final: set[str] = set()
            # 포함 페어는 반드시 동일 팀 유지 → 둘 다 잠금
            if seed is not None and seed % 3 == 0:
                locked_final.update({"이진원", "배연경"})
            # 제외 페어 멤버는 상대 팀과의 상대적 위치만 중요하므로 잠금까지는 불필요하지만, 요청 안정성 위해 포함
            locked_final.update({"노시현"})
            rebuild_ob_yb_exact(locked_final.union(exclude_names))
        except Exception:
            pass
    except Exception:
        pass

    return teams


def enforce_exclusion_pairs(teams: List[Team], pairs: List[Tuple[str, str]]) -> None:
    name_to_location: Dict[str, Tuple[int, bool, int]] = {}

    def locate(name: str) -> Optional[Tuple[int, bool, int]]:
        for i, t in enumerate(teams):
            if t.leader.name == name:
                return (i, True, -1)
            for idx, m in enumerate(t.members):
                if m.name == name:
                    return (i, False, idx)
        return None

    def team_has_name(team: Team, name: str) -> bool:
        if team.leader.name == name:
            return True
        return any(m.name == name for m in team.members)

    def try_resolve_pair(a: str, b: str) -> None:
        loc_a = locate(a)
        loc_b = locate(b)
        if not loc_a or not loc_b:
            return
        team_a, a_is_leader, a_idx = loc_a
        team_b, b_is_leader, b_idx = loc_b
        if team_a != team_b:
            return  # already satisfied

        conflicted_team = team_a
        # 우선 멤버 쪽을 이동(리더는 그대로 두는 것을 우선)
        # 후보 1: a가 멤버이면 a를 이동 시도
        if not a_is_leader:
            src_member = teams[conflicted_team].members[a_idx]
            # 같은 그룹 멤버와 스왑하여 팀별 그룹 수를 유지
            for j, t in enumerate(teams):
                if j == conflicted_team:
                    continue
                if team_has_name(t, b):
                    continue  # b가 있는 팀으로 보내지 않음
                for j_idx, other in enumerate(t.members):
                    if other.group == src_member.group:
                        teams[conflicted_team].members[a_idx], t.members[j_idx] = t.members[j_idx], teams[conflicted_team].members[a_idx]
                        return
        # 후보 2: b가 멤버이면 b를 이동 시도
        if not b_is_leader:
            src_member = teams[conflicted_team].members[b_idx]
            for j, t in enumerate(teams):
                if j == conflicted_team:
                    continue
                if team_has_name(t, a):
                    continue
                for j_idx, other in enumerate(t.members):
                    if other.group == src_member.group:
                        teams[conflicted_team].members[b_idx], t.members[j_idx] = t.members[j_idx], teams[conflicted_team].members[b_idx]
                        return
        # 그래도 불가하면 마지막 수단: 같은 그룹이 아니더라도 멤버와 스왑(팀 총원만 유지)
        if not a_is_leader:
            for j, t in enumerate(teams):
                if j == conflicted_team:
                    continue
                if team_has_name(t, b):
                    continue
                if t.members:
                    teams[conflicted_team].members[a_idx], t.members[0] = t.members[0], teams[conflicted_team].members[a_idx]
                    return
        if not b_is_leader:
            for j, t in enumerate(teams):
                if j == conflicted_team:
                    continue
                if team_has_name(t, a):
                    continue
                if t.members:
                    teams[conflicted_team].members[b_idx], t.members[0] = t.members[0], teams[conflicted_team].members[b_idx]
                    return

    for a, b in pairs:
        try_resolve_pair(a, b)


def enforce_inclusion_pair(teams: List[Team], pair: Tuple[str, str]) -> None:
    a, b = pair
    def locate(name: str) -> Optional[Tuple[int, bool, int]]:
        for i, t in enumerate(teams):
            if t.leader.name == name:
                return (i, True, -1)
            for idx, m in enumerate(t.members):
                if m.name == name:
                    return (i, False, idx)
        return None

    loc_a = locate(a)
    loc_b = locate(b)
    if not loc_a or not loc_b:
        return
    team_a, a_is_leader, a_idx = loc_a
    team_b, b_is_leader, b_idx = loc_b
    if team_a == team_b:
        return  # already same team

    # 스왑 전략: 우선 같은 그룹끼리 스왑을 시도하여 팀별 그룹 수 유지
    def swap_members(src_team: int, src_idx: int, dst_team: int) -> bool:
        src_member = teams[src_team].members[src_idx]
        for j_idx, other in enumerate(teams[dst_team].members):
            if other.group == src_member.group:
                teams[src_team].members[src_idx], teams[dst_team].members[j_idx] = teams[dst_team].members[j_idx], teams[src_team].members[src_idx]
                return True
        # 같은 그룹 후보가 없으면 첫 멤버와 교환(팀 총원만 유지)
        if teams[dst_team].members:
            teams[src_team].members[src_idx], teams[dst_team].members[0] = teams[dst_team].members[0], teams[src_team].members[src_idx]
            return True
        return False

    # 케이스 분기: 리더가 포함된 경우는 멤버 쪽만 이동
    if a_is_leader and not b_is_leader:
        swap_members(team_b, b_idx, team_a)
    elif b_is_leader and not a_is_leader:
        swap_members(team_a, a_idx, team_b)
    elif not a_is_leader and not b_is_leader:
        # 서로 다른 팀의 멤버 둘을 교차 스왑하여 같은 팀 만들기
        # 우선 a를 b의 팀으로 이동
        if not swap_members(team_a, a_idx, team_b):
            # 실패하면 b를 a의 팀으로 이동
            swap_members(team_b, b_idx, team_a)


def export_to_excel_bytes(teams: List[Team]) -> bytes:
    wb = Workbook()
    ws_by_team = wb.active
    ws_by_team.title = "ByTeam"
    ws_flat = wb.create_sheet("Flat")

    col = 1
    for team in teams:
        ws_by_team.cell(row=1, column=col, value=f"Team {team.index + 1}")
        ws_by_team.cell(row=2, column=col, value=f"Captain: {team.leader.name} ({team.leader.gender})")
        ws_by_team.cell(row=2, column=col + 1, value="captain")
        row = 3
        for m in team.members:
            ws_by_team.cell(row=row, column=col, value=m.name)
            ws_by_team.cell(row=row, column=col + 1, value=DISPLAY_LABELS.get(m.group, m.group))
            row += 1
        col += 3

    ws_flat.append(["team", "role", "group", "name", "gender"])
    for team in teams:
        ws_flat.append([team.index + 1, "captain", "captain", team.leader.name, team.leader.gender])
        for m in team.members:
            ws_flat.append([team.index + 1, "member", DISPLAY_LABELS.get(m.group, m.group), m.name, m.gender])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


def export_rooms_to_excel_bytes(male_rooms: List["Room"], female_rooms: List["Room"]) -> bytes:
    wb = Workbook()
    ws_m = wb.active
    ws_m.title = "Rooms_M"
    ws_f = wb.create_sheet("Rooms_F")
    ws_flat = wb.create_sheet("FlatRooms")

    # Rooms_M
    ws_m.append(["room", "name", "group", "gender"])
    for r in male_rooms:
        for mem in r.members:
            ws_m.append([r.index + 1, mem.name, mem.group, mem.gender])

    # Rooms_F
    ws_f.append(["room", "name", "group", "gender"])
    for r in female_rooms:
        for mem in r.members:
            ws_f.append([r.index + 1, mem.name, mem.group, mem.gender])

    # FlatRooms
    ws_flat.append(["gender_block", "room", "name", "group", "gender"])
    for r in male_rooms:
        for mem in r.members:
            ws_flat.append(["M", r.index + 1, mem.name, mem.group, mem.gender])
    for r in female_rooms:
        for mem in r.members:
            ws_flat.append(["F", r.index + 1, mem.name, mem.group, mem.gender])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


def read_default_or_upload(label: str, default_path: Path) -> bytes:
    uploaded = st.file_uploader(f"{label} CSV 업로드", type=["csv"], key=label)
    if uploaded is not None:
        return uploaded.read()
    with st.expander(f"{label} - 기본 파일 사용 경로 보기", expanded=False):
        st.code(str(default_path))
    return read_csv_from_disk(default_path)


def badge_html(group: str) -> str:
    color = GROUP_COLORS.get(group, "#666")
    return f'<span style="background:{color};color:#fff;padding:2px 8px;border-radius:12px;font-size:12px">{group}</span>'

# 전역 스타일(CSS)
st.markdown(
    """
    <style>
    html, body, .stApp{background:#ffffff !important; color:#0f172a !important}
    /* 상단 타이틀과 탭/버튼 대비 강화 */
    h1, h2, h3, h4, h5, h6 { color:#0f172a !important }
    .stApp header, .stApp [data-testid="stHeader"] { background:rgba(255,255,255,0.9) !important }
    .stTabs [data-baseweb="tab-list"] { background:#ffffff !important }
    .stTabs [data-baseweb="tab"], .stTabs [data-baseweb="tab"] p { color:#0f172a !important }
    .stButton>button { color:#0f172a !important; border-color:#cbd5e1 !important; background:#f8fafc !important }
    .stButton>button:hover { background:#f1f5f9 !important }
    /* 다운로드 버튼(엑셀) 가독성 강화 */
    [data-testid="baseButton-secondary"] button,
    .stDownloadButton>button {
        color:#0f172a !important; background:#f8fafc !important; border:1px solid #cbd5e1 !important;
    }
    [data-testid="baseButton-secondary"] button:hover,
    .stDownloadButton>button:hover { background:#f1f5f9 !important }
    /* 폼 레이블/설명 대비 강화 */
    label, .stMarkdown p, .stCaption, .stText, .stRadio, .stSelectbox, .stNumberInput, .stSlider, .stTextInput {
        color:#0f172a !important;
    }
    [data-testid="stTextInput"] label,
    [data-testid="stNumberInput"] label,
    [data-testid="stSlider"] label {
        color:#0f172a !important; font-weight:600 !important;
    }
    /* 입력창 텍스트/배경/플레이스홀더 */
    [data-testid="stTextInput"] input,
    [data-testid="stNumberInput"] input { color:#0f172a !important; background:#ffffff !important; border:1px solid #cbd5e1 !important; caret-color:#0f172a !important; }
    [data-testid="stTextInput"] input::placeholder,
    [data-testid="stNumberInput"] input::placeholder { color:#475569 !important; opacity:1 !important; }
    input[type="text"], input[type="number"] { background:#ffffff !important; color:#0f172a !important; border:1px solid #cbd5e1 !important; }
    [data-testid="stTextInput"] div:has(input), [data-testid="stNumberInput"] div:has(input) { background:transparent !important; }
    /* 슬라이더 대비 강화 */
    [data-testid="stSlider"] [role="slider"]{ background:#0f172a !important; border:2px solid #0f172a !important; }
    [data-testid="stSlider"] div[role="presentation"] div{ background:#e2e8f0 !important; }
    [data-testid="stSlider"] p{ color:#0f172a !important; }
    .team-card{background:#ffffff;border:1px solid #e5e7eb;border-radius:12px;padding:12px 14px;margin-bottom:10px;transition:border-color .2s, box-shadow .2s}
    .team-card.highlight{border-color:#f59e0b;box-shadow:0 0 0 3px rgba(245,158,11,.25);animation:glow .8s ease-in-out 2 alternate}
    .team-header{display:flex;justify-content:space-between;align-items:center;margin-bottom:8px}
    .team-title{font-weight:700;font-size:18px;white-space:nowrap;color:#0f172a !important}
    .leader-chip{background:transparent;border:none;color:#0f172a;border-radius:0;padding:0;margin:4px 0 8px;display:flex;align-items:center;gap:8px}
    .leader-crown{display:inline-flex;align-items:center;justify-content:center;font-size:18px;line-height:1}
    .member-list{display:flex;flex-direction:column;gap:6px}
    .member-item{display:flex;align-items:center;gap:8px}
    .member-name{font-size:16px;font-weight:400;color:#0f172a !important}
    .leader-name{font-size:20px;font-weight:800;color:#0f172a !important}
    .leader-label{display:inline-flex;align-items:center;justify-content:center;width:64px;padding:2px 8px;margin:0;border-radius:999px;font-size:12px;letter-spacing:.04em;text-transform:uppercase;background:#fde68a;border:1px solid #f59e0b;color:#78350f}
    .count-chip{background:#ffffff;border:1px solid #e5e7eb;border-radius:999px;padding:3px 10px;font-size:12px;margin-left:6px;color:#0f172a !important;white-space:nowrap;font-weight:600}
    .badge{display:inline-flex;align-items:center;justify-content:center;width:64px;padding:2px 8px;border-radius:999px;color:#fff;font-size:12px}
    .badge-ob{background:#1f77b4}.badge-yb{background:#2ca02c}.badge-girls{background:#d62728}.badge-leader{background:#7c3aed}
    @keyframes glow{from{box-shadow:0 0 0 0 rgba(245,158,11,.0)}to{box-shadow:0 0 0 6px rgba(245,158,11,.15)}}
    .spotlight{background:linear-gradient(135deg,#f0f9ff,#e9d5ff);border:1px solid #e5e7eb;border-radius:14px;padding:16px 18px;margin:8px 0;text-align:center;box-shadow:0 8px 24px rgba(15,23,42,.06)}
    .spotlight .label{font-size:12px;color:#64748b;text-transform:uppercase;letter-spacing:.08em;margin-bottom:6px}
    .spotlight strong{font-size:32px;color:#0f172a}
    /* 탭/헤더 내 텍스트 컬러 보정 */
    [data-testid="stMarkdownContainer"] p, [data-testid="stMarkdownContainer"] span { color:#0f172a !important; }
    </style>
    """,
    unsafe_allow_html=True,
)

def group_badge(group: str) -> str:
    label = DISPLAY_LABELS.get(group, group.upper())
    cls = f"badge badge-{group}"
    return f'<span class="{cls}">{label}</span>'

def member_item_html(mem: Member) -> str:
    return f'<div class="member-item">{group_badge(mem.group)}<span class="member-name">{mem.name}</span></div>'

def build_team_card_html(team_idx: int, leader: Member, members: List[Member]) -> str:
    total = 1 + len(members)
    ob_c = sum(1 for m in members if m.group == "ob")
    yb_c = sum(1 for m in members if m.group == "yb")
    g_c = sum(1 for m in members if m.group == "girls")
    members_html = "\n".join(member_item_html(m) for m in members)
    header_counts = (
        f'<span class="count-chip">총 {total}명</span>'
        f'<span class="count-chip">Veterans&nbsp;{ob_c}</span>'
        f'<span class="count-chip">Rookies&nbsp;{yb_c}</span>'
        f'<span class="count-chip">Girls&nbsp;{g_c}</span>'
    )
    return (
        f'<div class="team-card">'
        f'  <div class="team-header">'
        f'    <div class="team-title">Team&nbsp;{team_idx + 1}</div>'
        f'    <div>{header_counts}</div>'
        f'  </div>'
        f'  <div class="leader-chip"><span class="leader-label">CAPTAIN</span><span class="leader-name">{leader.name}</span><span class="leader-crown"> 👑</span></div>'
        f'  <div class="member-list">{members_html}</div>'
        f'</div>'
    )


def build_room_card_html(room_idx: int, members: List[Member], title_prefix: str = "Room") -> str:
    members_html = "\n".join(member_item_html(m) for m in members)
    return (
        f'<div class="team-card">'
        f'  <div class="team-header">'
        f'    <div class="team-title">{title_prefix} {room_idx + 1}</div>'
        f'  </div>'
        f'  <div class="member-list">{members_html}</div>'
        f'</div>'
    )


def assign_rooms(
    leaders: List[Member],
    ob_list: List[Member],
    yb_list: List[Member],
    girls_list: List[Member],
    room_size: int = 4,
    seed: Optional[int] = None,
) -> Tuple[List[Room], List[Room]]:
    if seed is None:
        seed = int(time.time() * 1000) % (2**32 - 1)
    rng = random.Random(seed)

    male_leaders = [m for m in leaders if m.gender == "M"]
    female_leaders = [m for m in leaders if m.gender == "F"]

    # 중복 제거: 리더 우선, 그 다음 기존 그룹
    def dedup_by_name(preferred: List[Member], others: List[Member]) -> List[Member]:
        name_to_member: Dict[str, Member] = {}
        for m in preferred:
            if m.name.strip():
                name_to_member.setdefault(m.name, m)
        for m in others:
            if m.name.strip() and m.name not in name_to_member:
                name_to_member[m.name] = m
        return list(name_to_member.values())

    male_pool = dedup_by_name(male_leaders, ob_list + yb_list)
    female_pool = dedup_by_name(female_leaders, girls_list)

    rng.shuffle(male_pool)
    rng.shuffle(female_pool)

    def build_room_sizes(total: int, preferred_size: int) -> List[int]:
        if total <= 0:
            return []
        if total < 3:
            return [total]
        k = total // preferred_size
        r = total % preferred_size
        sizes: List[int] = []
        if r == 0:
            sizes = [preferred_size] * k
        elif r == 3:
            sizes = [preferred_size] * k + [3]
        elif r == 2:
            if k >= 1:
                sizes = [preferred_size] * (k - 1) + [3, 3]
            else:
                sizes = [2]
        elif r == 1:
            if k >= 2:
                sizes = [preferred_size] * (k - 2) + [3, 3, 3]
            elif k == 1:
                sizes = [3, 2]
            else:
                sizes = [1]
        return sizes

    def compose_rooms(members: List[Member], preferred_size: int) -> List[Room]:
        sizes = build_room_sizes(len(members), preferred_size)
        rooms: List[Room] = []
        cursor = 0
        for sz in sizes:
            rooms.append(Room(index=len(rooms), members=members[cursor: cursor + sz]))
            cursor += sz
        return rooms

    return compose_rooms(male_pool, room_size), compose_rooms(female_pool, room_size)


st.title("조 추첨기")

# 세션 상태 초기화
for _k in [
    "leaders_bytes",
    "ob_bytes",
    "yb_bytes",
    "girls_bytes",
    "seed_str",
    "teams_result",
    "rooms_result_m",
    "rooms_result_f",
    "rooms_seed_str",
    "rooms_reveal_pending",
]:
    st.session_state.setdefault(_k, None)

tab_settings, tab_draw, tab_rooms = st.tabs(["설정", "조 추첨", "룸메이트"]) 

with tab_settings:
    st.subheader("설정")
    st.caption("CSV는 UTF-8 인코딩 권장. 헤더: leaders=name,gender / ob,yb,girls=name")
    st.session_state["seed_str"] = st.text_input(
        "Seed (선택)", value=st.session_state.get("seed_str") or ""
    )
    st.session_state["rooms_seed_str"] = st.text_input(
        "Room Seed (선택)", value=st.session_state.get("rooms_seed_str") or ""
    )
    col_t1, col_t2 = st.columns(2)
    with col_t1:
        st.session_state["highlight_sec"] = st.slider(
            "하이라이트 시간(초)", min_value=0.0, max_value=1.0,
            value=float(st.session_state.get("highlight_sec") or 0.15), step=0.01
        )
    with col_t2:
        st.session_state["interval_sec"] = st.slider(
            "노출 텀(초)", min_value=0.0, max_value=1.0,
            value=float(st.session_state.get("interval_sec") or 0.24), step=0.01
        )
    st.divider()
    st.write("파일 업로드 또는 기본 파일 사용")
    leaders_bytes = read_default_or_upload("Captains", DATA_DIR / "leaders.csv")
    ob_bytes = read_default_or_upload("Veterans", DATA_DIR / "ob.csv")
    yb_bytes = read_default_or_upload("Rookies", DATA_DIR / "yb.csv")
    girls_bytes = read_default_or_upload("Girls", DATA_DIR / "girls.csv")

    st.session_state["leaders_bytes"] = leaders_bytes
    st.session_state["ob_bytes"] = ob_bytes
    st.session_state["yb_bytes"] = yb_bytes
    st.session_state["girls_bytes"] = girls_bytes

    try:
        leaders_preview = read_leaders_csv_from_bytes(leaders_bytes) if leaders_bytes else []
        ob_preview = read_names_csv_from_bytes(ob_bytes, group="ob", gender="M") if ob_bytes else []
        yb_preview = read_names_csv_from_bytes(yb_bytes, group="yb", gender="M") if yb_bytes else []
        girls_preview = read_names_csv_from_bytes(girls_bytes, group="girls", gender="F") if girls_bytes else []
        st.info(
            f"캡틴 {len(leaders_preview)}명, Veterans {len(ob_preview)}명, Rookies {len(yb_preview)}명, Girls {len(girls_preview)}명"
        )
    except Exception as e:
        st.warning(str(e))

with tab_draw:
    st.subheader("조 추첨")
    status_ph = st.empty()
    if st.button("추첨 실행", type="primary"):
        try:
            status_ph.info("추첨 중…")
            leaders_bytes = st.session_state.get("leaders_bytes") or read_csv_from_disk(DATA_DIR / "leaders.csv")
            ob_bytes = st.session_state.get("ob_bytes") or read_csv_from_disk(DATA_DIR / "ob.csv")
            yb_bytes = st.session_state.get("yb_bytes") or read_csv_from_disk(DATA_DIR / "yb.csv")
            girls_bytes = st.session_state.get("girls_bytes") or read_csv_from_disk(DATA_DIR / "girls.csv")

            leaders = read_leaders_csv_from_bytes(leaders_bytes)
            ob_list = read_names_csv_from_bytes(ob_bytes, group="ob", gender="M")
            yb_list = read_names_csv_from_bytes(yb_bytes, group="yb", gender="M")
            girls_list = read_names_csv_from_bytes(girls_bytes, group="girls", gender="F")

            leader_names = {m.name for m in leaders}
            before_counts = (len(ob_list), len(yb_list), len(girls_list))
            ob_list = [m for m in ob_list if m.name not in leader_names]
            yb_list = [m for m in yb_list if m.name not in leader_names]
            girls_list = [m for m in girls_list if m.name not in leader_names]
            after_counts = (len(ob_list), len(yb_list), len(girls_list))
            if before_counts != after_counts:
                removed_ob = before_counts[0] - after_counts[0]
                removed_yb = before_counts[1] - after_counts[1]
                removed_girls = before_counts[2] - after_counts[2]
                st.info(
                    f"캡틴과 중복된 이름을 제외했습니다. 제거됨 - Veterans:{removed_ob}, Rookies:{removed_yb}, Girls:{removed_girls}"
                )

            seed_str = st.session_state.get("seed_str") or ""
            seed_val: Optional[int] = int(seed_str) if seed_str.strip() else None
            teams = assign_members_to_teams(leaders, ob_list, yb_list, girls_list, seed=seed_val)

            st.session_state["teams_result"] = teams
            st.session_state["reveal_pending"] = True
            status_ph.info("추첨 중…")
        except Exception as e:
            status_ph.empty()
            st.error(str(e))

    teams_draw = st.session_state.get("teams_result")
    if teams_draw:
        st.divider()
        # 전광판(상단)
        spotlight = st.empty()
        # 팀 카드 자리(리더만 먼저 표시)
        team_placeholders: List[st.delta_generator.DeltaGenerator] = []
        for row_start in (0, 4):
            cols = st.columns(4)
            for j in range(4):
                with cols[j]:
                    ph = st.empty()
                    team_placeholders.append(ph)
        for i in range(8):
            team_placeholders[i].markdown(
                build_team_card_html(teams_draw[i].index, teams_draw[i].leader, []),
                unsafe_allow_html=True,
            )

        # 전광판 + 순차 공개(사용자 설정 속도)
        highlight_sec = float(st.session_state.get("highlight_sec") or 0.15)
        interval_sec = float(st.session_state.get("interval_sec") or 0.24)
        if st.session_state.get("reveal_pending"):
            status_ph.info("추첨 진행 중…")
            revealed_by_team: List[List[Member]] = [[] for _ in range(8)]
            seed_txt = (st.session_state.get("seed_str") or "").strip()
            try:
                seed_val_for_reveal = int(seed_txt) if seed_txt else None
            except Exception:
                seed_val_for_reveal = None
            rng = random.Random(seed_val_for_reveal if seed_val_for_reveal is not None else int(time.time()))
            reveal_order: List[Tuple[int, Member]] = []
            # 공개 순서: OB → YB → Girls (각 그룹 내부는 무작위)
            for category in ["ob", "yb", "girls"]:
                for i, t in enumerate(teams_draw):
                    group_members = [m for m in t.members if m.group == category]
                    rng.shuffle(group_members)
                    for m in group_members:
                        reveal_order.append((i, m))

            for _, (ti, mem) in enumerate(reveal_order):
                spotlight.markdown(
                    f"<div class='spotlight'><div class='label'>Who's next?</div><strong>{mem.name}</strong></div>",
                    unsafe_allow_html=True,
                )
                revealed_by_team[ti].append(mem)
                # 하이라이트 효과로 짧게 반짝
                html_temp = build_team_card_html(teams_draw[ti].index, teams_draw[ti].leader, revealed_by_team[ti])
                html_temp = html_temp.replace("team-card", "team-card highlight", 1)
                team_placeholders[ti].markdown(html_temp, unsafe_allow_html=True)
                time.sleep(max(0.0, highlight_sec))
                team_placeholders[ti].markdown(
                    build_team_card_html(teams_draw[ti].index, teams_draw[ti].leader, revealed_by_team[ti]),
                    unsafe_allow_html=True,
                )
                # 전체 속도 조절(사용자 설정)
                time.sleep(max(0.0, interval_sec))
            spotlight.empty()
            status_ph.success("추첨 완료!")
            st.session_state["reveal_pending"] = False
        else:
            # 애니메이션 없이 전체 멤버 즉시 렌더
            for i in range(8):
                team_placeholders[i].markdown(
                    build_team_card_html(teams_draw[i].index, teams_draw[i].leader, teams_draw[i].members),
                    unsafe_allow_html=True,
                )
            status_ph.success("추첨 완료!")

        # 요약 통계
        st.divider()
        cols_stat = st.columns(4)
        for i, col in enumerate(cols_stat):
            t = teams_draw[i]
            with col:
                st.caption(
                    f"Team {t.index + 1}: 총 {1 + len(t.members)}명 (Veterans:{sum(1 for m in t.members if m.group=='ob')}, Rookies:{sum(1 for m in t.members if m.group=='yb')}, Girls:{sum(1 for m in t.members if m.group=='girls')})"
                )
        cols_stat2 = st.columns(4)
        for i, col in enumerate(cols_stat2):
            t = teams_draw[4 + i]
            with col:
                st.caption(
                    f"Team {t.index + 1}: 총 {1 + len(t.members)}명 (Veterans:{sum(1 for m in t.members if m.group=='ob')}, Rookies:{sum(1 for m in t.members if m.group=='yb')}, Girls:{sum(1 for m in t.members if m.group=='girls')})"
                )

        # 엑셀 다운로드
        xlsx_bytes = export_to_excel_bytes(teams_draw)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        st.download_button(
            label="엑셀 다운로드",
            data=xlsx_bytes,
            file_name=f"draw_result_{ts}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_button_draw",
        )
        st.balloons()


with tab_rooms:
    st.subheader("룸메이트 배정 (4인 1실, 성별 분리)")
    st.caption("캡틴 여부와 무관하게 성별을 기준으로 4인실로 배정합니다. 업로드한 CSV(캡틴/Veterans/Rookies/Girls)를 그대로 재사용합니다.")

    # 옵션: 방 크기(기본 4). 시드는 설정 탭(Room Seed) 사용
    room_size = st.number_input("방 정원", min_value=2, max_value=6, value=4, step=1)

    col_rooms_actions = st.columns(2)
    status_rooms = st.empty()
    with col_rooms_actions[0]:
        if st.button("룸메이트 배정 실행", type="primary"):
            try:
                leaders_bytes = st.session_state.get("leaders_bytes") or read_csv_from_disk(DATA_DIR / "leaders.csv")
                ob_bytes = st.session_state.get("ob_bytes") or read_csv_from_disk(DATA_DIR / "ob.csv")
                yb_bytes = st.session_state.get("yb_bytes") or read_csv_from_disk(DATA_DIR / "yb.csv")
                girls_bytes = st.session_state.get("girls_bytes") or read_csv_from_disk(DATA_DIR / "girls.csv")

                leaders_all = read_leaders_csv_from_bytes(leaders_bytes) if leaders_bytes else []
                ob_list = read_names_csv_from_bytes(ob_bytes, group="ob", gender="M") if ob_bytes else []
                yb_list = read_names_csv_from_bytes(yb_bytes, group="yb", gender="M") if yb_bytes else []
                girls_list = read_names_csv_from_bytes(girls_bytes, group="girls", gender="F") if girls_bytes else []

                # 시드 처리: 설정 탭의 Room Seed 사용
                seed_val_rooms: Optional[int] = None
                seed_text = (st.session_state.get("rooms_seed_str") or "").strip()
                if seed_text:
                    try:
                        seed_val_rooms = int(seed_text)
                    except Exception:
                        seed_val_rooms = None

                rooms_m, rooms_f = assign_rooms(
                    leaders=leaders_all,
                    ob_list=ob_list,
                    yb_list=yb_list,
                    girls_list=girls_list,
                    room_size=int(room_size),
                    seed=seed_val_rooms,
                )
                st.session_state["rooms_result_m"] = rooms_m
                st.session_state["rooms_result_f"] = rooms_f
                st.session_state["rooms_reveal_pending"] = True
                status_rooms.info("룸메이트 배정 중…")
            except Exception as e:
                status_rooms.error(str(e))

    rooms_m = st.session_state.get("rooms_result_m")
    rooms_f = st.session_state.get("rooms_result_f")

    if rooms_m or rooms_f:
        st.divider()
        spotlight_rooms = st.empty()

        # 남자 방 자리(빈 카드 먼저)
        male_room_placeholders: List[st.delta_generator.DeltaGenerator] = []
        if rooms_m:
            st.markdown("### 남자 방")
            cols_m = st.columns(4)
            for i, r in enumerate(rooms_m):
                with cols_m[i % 4]:
                    ph = st.empty()
                    male_room_placeholders.append(ph)
            for i, r in enumerate(rooms_m):
                male_room_placeholders[i].markdown(
                    build_room_card_html(r.index, [], title_prefix="Room(M)"),
                    unsafe_allow_html=True,
                )

        # 여자 방 자리(빈 카드 먼저)
        female_room_placeholders: List[st.delta_generator.DeltaGenerator] = []
        if rooms_f:
            st.markdown("### 여자 방")
            cols_f = st.columns(4)
            for i, r in enumerate(rooms_f):
                with cols_f[i % 4]:
                    ph = st.empty()
                    female_room_placeholders.append(ph)
            for i, r in enumerate(rooms_f):
                female_room_placeholders[i].markdown(
                    build_room_card_html(r.index, [], title_prefix="Room(F)"),
                    unsafe_allow_html=True,
                )

        highlight_sec = float(st.session_state.get("highlight_sec") or 0.15)
        interval_sec = float(st.session_state.get("interval_sec") or 0.24)

        if st.session_state.get("rooms_reveal_pending"):
            # 전개 순서(M/F 섞어서)
            reveal_order: List[Tuple[str, int, Member]] = []
            if rooms_m:
                for i, r in enumerate(rooms_m):
                    for m in r.members:
                        reveal_order.append(("M", i, m))
            if rooms_f:
                for i, r in enumerate(rooms_f):
                    for m in r.members:
                        reveal_order.append(("F", i, m))

            # 시드 기반 셔플(설정의 Room Seed)
            seed_text = (st.session_state.get("rooms_seed_str") or "").strip()
            try:
                seed_val_for_reveal = int(seed_text) if seed_text else None
            except Exception:
                seed_val_for_reveal = None
            rng = random.Random(seed_val_for_reveal if seed_val_for_reveal is not None else int(time.time()))
            rng.shuffle(reveal_order)

            revealed_m: List[List[Member]] = [[] for _ in range(len(rooms_m or []))]
            revealed_f: List[List[Member]] = [[] for _ in range(len(rooms_f or []))]

            for _, (gender_block, idx, mem) in enumerate(reveal_order):
                spotlight_rooms.markdown(
                    f"<div class='spotlight'><div class='label'>Who's next?</div><strong>{mem.name}</strong></div>",
                    unsafe_allow_html=True,
                )
                if gender_block == "M" and rooms_m:
                    revealed_m[idx].append(mem)
                    html_temp = build_room_card_html(rooms_m[idx].index, revealed_m[idx], title_prefix="Room(M)")
                    html_temp = html_temp.replace("team-card", "team-card highlight", 1)
                    male_room_placeholders[idx].markdown(html_temp, unsafe_allow_html=True)
                    time.sleep(max(0.0, highlight_sec))
                    male_room_placeholders[idx].markdown(
                        build_room_card_html(rooms_m[idx].index, revealed_m[idx], title_prefix="Room(M)"),
                        unsafe_allow_html=True,
                    )
                elif gender_block == "F" and rooms_f:
                    revealed_f[idx].append(mem)
                    html_temp = build_room_card_html(rooms_f[idx].index, revealed_f[idx], title_prefix="Room(F)")
                    html_temp = html_temp.replace("team-card", "team-card highlight", 1)
                    female_room_placeholders[idx].markdown(html_temp, unsafe_allow_html=True)
                    time.sleep(max(0.0, highlight_sec))
                    female_room_placeholders[idx].markdown(
                        build_room_card_html(rooms_f[idx].index, revealed_f[idx], title_prefix="Room(F)"),
                        unsafe_allow_html=True,
                    )
                time.sleep(max(0.0, interval_sec))

            spotlight_rooms.empty()
            status_rooms.success("룸메이트 배정 완료")
            st.session_state["rooms_reveal_pending"] = False
        else:
            # 애니메이션 없이 전체 렌더
            if rooms_m:
                for i, r in enumerate(rooms_m):
                    male_room_placeholders[i].markdown(
                        build_room_card_html(r.index, r.members, title_prefix="Room(M)"),
                        unsafe_allow_html=True,
                    )
            if rooms_f:
                for i, r in enumerate(rooms_f):
                    female_room_placeholders[i].markdown(
                        build_room_card_html(r.index, r.members, title_prefix="Room(F)"),
                        unsafe_allow_html=True,
                    )

        # 요약
        st.divider()
        if rooms_m:
            total_m = sum(len(r.members) for r in rooms_m)
            st.caption(f"남자: 방 {len(rooms_m)}개, 총 {total_m}명")
        if rooms_f:
            total_f = sum(len(r.members) for r in rooms_f)
            st.caption(f"여자: 방 {len(rooms_f)}개, 총 {total_f}명")

        # 엑셀 다운로드
        xlsx_rooms = export_rooms_to_excel_bytes(rooms_m or [], rooms_f or [])
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        st.download_button(
            label="룸메이트 엑셀 다운로드",
            data=xlsx_rooms,
            file_name=f"roommates_{ts}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_button_rooms",
        )
