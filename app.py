import csv
import os
import random
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook


DATA_DIR = Path(__file__).parent / "data"
OUTPUT_DIR = Path(__file__).parent / "output"


@dataclass
class Member:
    name: str
    group: str  # one of {"leader", "ob", "yb", "girls"}
    gender: str  # one of {"M", "F"}


@dataclass
class Team:
    index: int
    leader: Member
    members: List[Member] = field(default_factory=list)

    def add_member(self, member: Member) -> None:
        self.members.append(member)

    def all_people(self) -> List[Member]:
        return [self.leader] + self.members


def read_leaders_csv(path: Path) -> List[Member]:
    leaders: List[Member] = []
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = (row.get("name") or row.get("Name") or "").strip()
            gender = (row.get("gender") or row.get("Gender") or "").strip().upper()
            if not name:
                continue
            if gender not in {"M", "F"}:
                raise ValueError(f"리더 성별은 M/F 로 표기해야 합니다: {name} -> {gender}")
            leaders.append(Member(name=name, group="leader", gender=gender))
    return leaders


def read_names_csv(path: Path, group: str, gender: str) -> List[Member]:
    members: List[Member] = []
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = (row.get("name") or row.get("Name") or "").strip()
            if not name:
                continue
            members.append(Member(name=name, group=group, gender=gender))
    return members


def compute_group_targets(
    leaders: List[Member],
    ob_count: int,
    yb_count: int,
    girls_count: int,
) -> Tuple[Dict[int, Dict[str, int]], List[int], List[int]]:
    """
    반환값:
      - targets[team_index] = {"ob": x, "yb": y, "girls": z}
      - male_leader_team_indices
      - female_leader_team_indices
    분배 규칙:
      - ob/yb(남) 잔여 인원은 여리더 팀부터 우선 1명씩 추가
      - girls(여) 잔여 인원은 남리더 팀부터 우선 1명씩 추가
    """
    num_teams = len(leaders)
    male_leader_team_indices = [i for i, ld in enumerate(leaders) if ld.gender == "M"]
    female_leader_team_indices = [i for i, ld in enumerate(leaders) if ld.gender == "F"]

    def base_and_remainder(total: int) -> Tuple[int, int]:
        return total // num_teams, total % num_teams

    base_ob, rem_ob = base_and_remainder(ob_count)
    base_yb, rem_yb = base_and_remainder(yb_count)
    base_girls, rem_girls = base_and_remainder(girls_count)

    targets: Dict[int, Dict[str, int]] = {
        i: {"ob": base_ob, "yb": base_yb, "girls": base_girls} for i in range(num_teams)
    }

    # 남성 그룹(ob, yb) 잔여는 여리더 팀부터
    for idx in female_leader_team_indices[:rem_ob]:
        targets[idx]["ob"] += 1
    for idx in female_leader_team_indices[rem_ob:rem_ob]:
        pass  # no-op, for readability
    for idx in female_leader_team_indices:
        pass

    # 나머지 ob 잔여가 여리더 팀 수보다 많을 수 있어, 전체 팀에 라운드로빈으로 분배
    if rem_ob > len(female_leader_team_indices):
        remaining = rem_ob - len(female_leader_team_indices)
        others = [i for i in range(num_teams) if i not in female_leader_team_indices]
        for i in range(remaining):
            targets[others[i % len(others)]]['ob'] += 1

    # yb 도 동일 규칙
    for idx in female_leader_team_indices[:rem_yb]:
        targets[idx]["yb"] += 1
    if rem_yb > len(female_leader_team_indices):
        remaining = rem_yb - len(female_leader_team_indices)
        others = [i for i in range(num_teams) if i not in female_leader_team_indices]
        for i in range(remaining):
            targets[others[i % len(others)]]['yb'] += 1

    # 여성 그룹(girls) 잔여는 남리더 팀부터
    for idx in male_leader_team_indices[:rem_girls]:
        targets[idx]["girls"] += 1
    if rem_girls > len(male_leader_team_indices):
        remaining = rem_girls - len(male_leader_team_indices)
        others = [i for i in range(num_teams) if i not in male_leader_team_indices]
        for i in range(remaining):
            targets[others[i % len(others)]]['girls'] += 1

    return targets, male_leader_team_indices, female_leader_team_indices


def assign_members_to_teams(
    leaders: List[Member],
    ob_list: List[Member],
    yb_list: List[Member],
    girls_list: List[Member],
    seed: Optional[int] = None,
) -> List[Team]:
    if seed is None:
        seed = int(time.time() * 1000) % (2**32 - 1)
    rng = random.Random(seed)

    num_teams = len(leaders)
    if num_teams != 8:
        raise ValueError("리더 수는 반드시 8명이어야 합니다.")
    male_count = sum(1 for m in leaders if m.gender == "M")
    female_count = sum(1 for m in leaders if m.gender == "F")
    if male_count != 4 or female_count != 4:
        raise ValueError("리더 성별은 남 4, 여 4 이어야 합니다.")

    targets, _, _ = compute_group_targets(
        leaders, ob_count=len(ob_list), yb_count=len(yb_list), girls_count=len(girls_list)
    )

    rng.shuffle(ob_list)
    rng.shuffle(yb_list)
    rng.shuffle(girls_list)

    teams = [Team(index=i, leader=leaders[i]) for i in range(num_teams)]

    def pop_many(src: List[Member], count: int) -> List[Member]:
        if count <= 0:
            return []
        if count > len(src):
            raise ValueError("분배 대상 인원이 부족합니다. 입력 CSV를 확인하세요.")
        result = src[:count]
        del src[:count]
        return result

    # 그룹별 목표치만큼 순차 분배
    for i in range(num_teams):
        teams[i].members.extend(pop_many(ob_list, targets[i]["ob"]))
    for i in range(num_teams):
        teams[i].members.extend(pop_many(yb_list, targets[i]["yb"]))
    for i in range(num_teams):
        teams[i].members.extend(pop_many(girls_list, targets[i]["girls"]))

    # 팀별 내부 섞기(보기 좋게)
    for team in teams:
        rng.shuffle(team.members)

    return teams


def export_to_excel(teams: List[Team], out_dir: Path) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = out_dir / f"draw_result_{timestamp}.xlsx"

    wb = Workbook()
    ws_by_team = wb.active
    ws_by_team.title = "ByTeam"
    ws_flat = wb.create_sheet("Flat")

    # ByTeam: 각 팀을 두 컬럼으로 구성 (이름 / 그룹)
    col = 1
    for team in teams:
        ws_by_team.cell(row=1, column=col, value=f"Team {team.index + 1}")
        ws_by_team.cell(
            row=2,
            column=col,
            value=f"Leader: {team.leader.name} ({team.leader.gender})",
        )
        ws_by_team.cell(row=2, column=col + 1, value="leader")
        row = 3
        for m in team.members:
            ws_by_team.cell(row=row, column=col, value=m.name)
            ws_by_team.cell(row=row, column=col + 1, value=m.group)
            row += 1
        col += 3

    # Flat 시트
    ws_flat.append(["team", "role", "group", "name", "gender"])
    for team in teams:
        ws_flat.append([
            team.index + 1,
            "leader",
            "leader",
            team.leader.name,
            team.leader.gender,
        ])
        for m in team.members:
            ws_flat.append([team.index + 1, "member", m.group, m.name, m.gender])

    wb.save(out_path)
    return out_path


class TeamDrawerApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("조 추첨기")
        self.root.geometry("1200x750")

        self.seed_var = tk.StringVar()
        self.leaders_path = tk.StringVar(value=str(DATA_DIR / "leaders.csv"))
        self.ob_path = tk.StringVar(value=str(DATA_DIR / "ob.csv"))
        self.yb_path = tk.StringVar(value=str(DATA_DIR / "yb.csv"))
        self.girls_path = tk.StringVar(value=str(DATA_DIR / "girls.csv"))

        self.leaders: List[Member] = []
        self.ob_list: List[Member] = []
        self.yb_list: List[Member] = []
        self.girls_list: List[Member] = []
        self.teams: List[Team] = []

        self._build_ui()

    def _build_ui(self) -> None:
        top = ttk.Frame(self.root)
        top.pack(fill=tk.X, padx=12, pady=8)

        ttk.Label(top, text="Seed").grid(row=0, column=0, sticky=tk.W)
        ttk.Entry(top, textvariable=self.seed_var, width=12).grid(row=0, column=1, padx=4)

        def add_file_row(r: int, label: str, var: tk.StringVar):
            ttk.Label(top, text=label).grid(row=r, column=2, sticky=tk.W, padx=(16, 4))
            ttk.Entry(top, textvariable=var, width=60).grid(row=r, column=3, padx=4)
            ttk.Button(
                top,
                text="찾기",
                command=lambda: var.set(
                    filedialog.askopenfilename(
                        title=f"{label} CSV 선택",
                        filetypes=[("CSV Files", "*.csv")],
                        initialdir=str(DATA_DIR),
                    )
                ),
            ).grid(row=r, column=4, padx=4)

        add_file_row(0, "Leaders", self.leaders_path)
        add_file_row(1, "OB", self.ob_path)
        add_file_row(2, "YB", self.yb_path)
        add_file_row(3, "Girls", self.girls_path)

        btns = ttk.Frame(self.root)
        btns.pack(fill=tk.X, padx=12, pady=8)
        ttk.Button(btns, text="데이터 불러오기", command=self.load_data).pack(side=tk.LEFT)
        ttk.Button(btns, text="추첨 시작", command=self.start_draw).pack(side=tk.LEFT, padx=8)
        ttk.Button(btns, text="엑셀로 저장", command=self.save_excel).pack(side=tk.LEFT, padx=8)
        ttk.Button(btns, text="초기화", command=self.reset).pack(side=tk.LEFT, padx=8)

        self.status_var = tk.StringVar(value="데이터를 불러와 주세요.")
        ttk.Label(self.root, textvariable=self.status_var).pack(anchor=tk.W, padx=12)

        self.board = ttk.Frame(self.root)
        self.board.pack(fill=tk.BOTH, expand=True, padx=12, pady=8)
        self.team_frames: List[tk.Listbox] = []
        self._build_board()

        self.roulette_var = tk.StringVar()
        self.roulette_label = ttk.Label(self.root, textvariable=self.roulette_var, font=("Arial", 20))
        self.roulette_label.pack(pady=(0, 10))

    def _build_board(self) -> None:
        for child in self.board.winfo_children():
            child.destroy()
        self.team_frames.clear()
        grid = ttk.Frame(self.board)
        grid.pack(fill=tk.BOTH, expand=True)
        for i in range(8):
            frame = ttk.Frame(grid, relief=tk.RIDGE, borderwidth=1)
            frame.grid(row=i // 4, column=i % 4, sticky=tk.NSEW, padx=6, pady=6)
            grid.columnconfigure(i % 4, weight=1)
            grid.rowconfigure(i // 4, weight=1)
            ttk.Label(frame, text=f"Team {i + 1}").pack()
            lb = tk.Listbox(frame, height=12)
            lb.pack(fill=tk.BOTH, expand=True)
            self.team_frames.append(lb)

    def load_data(self) -> None:
        try:
            leaders_path = Path(self.leaders_path.get())
            ob_path = Path(self.ob_path.get())
            yb_path = Path(self.yb_path.get())
            girls_path = Path(self.girls_path.get())

            if not leaders_path.exists():
                raise FileNotFoundError(f"리더 파일이 없습니다: {leaders_path}")

            self.leaders = read_leaders_csv(leaders_path)
            self.ob_list = read_names_csv(ob_path, group="ob", gender="M") if ob_path.exists() else []
            self.yb_list = read_names_csv(yb_path, group="yb", gender="M") if yb_path.exists() else []
            self.girls_list = (
                read_names_csv(girls_path, group="girls", gender="F") if girls_path.exists() else []
            )

            self.status_var.set(
                f"불러오기 완료 - Leaders: {len(self.leaders)}, OB: {len(self.ob_list)}, YB: {len(self.yb_list)}, Girls: {len(self.girls_list)}"
            )

            self.refresh_board(leaders_only=True)
        except Exception as e:
            messagebox.showerror("에러", str(e))

    def refresh_board(self, leaders_only: bool = False) -> None:
        for i, lb in enumerate(self.team_frames):
            lb.delete(0, tk.END)
            if i < len(self.leaders):
                ld = self.leaders[i]
                lb.insert(tk.END, f"[LEADER] {ld.name} ({ld.gender})")
            if not leaders_only and i < len(self.teams):
                for m in self.teams[i].members:
                    lb.insert(tk.END, f"- {m.name} [{m.group}]")

    def start_draw(self) -> None:
        if not self.leaders:
            try:
                self.load_data()
            except Exception:
                return
        try:
            seed_val: Optional[int] = None
            seed_str = self.seed_var.get().strip()
            if seed_str:
                seed_val = int(seed_str)

            # 실제 배정 계산은 먼저 수행
            planned_teams = assign_members_to_teams(
                self.leaders[:], self.ob_list[:], self.yb_list[:], self.girls_list[:], seed=seed_val
            )

            # 애니메이션을 위한 큐 구성
            assignment_queue: List[Tuple[int, Member]] = []
            for team in planned_teams:
                for m in team.members:
                    assignment_queue.append((team.index, m))
            rng = random.Random(seed_val or int(time.time()))
            rng.shuffle(assignment_queue)

            # 애니메이션 상태
            self.teams = [Team(index=i, leader=self.leaders[i]) for i in range(len(self.leaders))]
            self.refresh_board(leaders_only=True)

            self.status_var.set("추첨 중...")

            def animate_next():
                if not assignment_queue:
                    self.status_var.set("추첨 완료!")
                    self.refresh_board(leaders_only=False)
                    return

                team_idx, member = assignment_queue.pop()
                # 룰렛처럼 이름 굴리기
                spin_names = [member.name]
                # 남은 인원 중 일부 랜덤 샘플
                for _ in range(5):
                    spin_names.append(member.name)

                def spin(i: int):
                    if i >= len(spin_names):
                        # 확정 배치
                        self.teams[team_idx].add_member(member)
                        self.team_frames[team_idx].insert(
                            tk.END, f"- {member.name} [{member.group}]"
                        )
                        self.root.after(150, animate_next)
                        return
                    self.roulette_var.set(f"{spin_names[i]}")
                    self.root.after(70, lambda: spin(i + 1))

                spin(0)

            animate_next()

        except Exception as e:
            messagebox.showerror("에러", str(e))

    def save_excel(self) -> None:
        try:
            if not self.teams:
                messagebox.showinfo("안내", "추첨 결과가 없습니다. 먼저 추첨을 진행하세요.")
                return
            out_path = export_to_excel(self.teams, OUTPUT_DIR)
            messagebox.showinfo("완료", f"엑셀 저장 완료:\n{out_path}")
        except Exception as e:
            messagebox.showerror("에러", str(e))

    def reset(self) -> None:
        self.teams = []
        self._build_board()
        if self.leaders:
            self.refresh_board(leaders_only=True)
        self.status_var.set("초기화되었습니다.")


def ensure_templates() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # 템플릿 파일 생성(없을 때만)
    templates = {
        DATA_DIR / "leaders.csv": [
            ["name", "gender"],
            ["리더1", "M"],
            ["리더2", "M"],
            ["리더3", "M"],
            ["리더4", "M"],
            ["리더5", "F"],
            ["리더6", "F"],
            ["리더7", "F"],
            ["리더8", "F"],
        ],
        DATA_DIR / "ob.csv": [["name"], ["OB1"], ["OB2"], ["OB3"], ["OB4"]],
        DATA_DIR / "yb.csv": [["name"], ["YB1"], ["YB2"], ["YB3"], ["YB4"]],
        DATA_DIR / "girls.csv": [["name"], ["G1"], ["G2"], ["G3"], ["G4"]],
    }
    for path, rows in templates.items():
        if not path.exists():
            with path.open("w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerows(rows)


def main() -> None:
    ensure_templates()
    root = tk.Tk()
    app = TeamDrawerApp(root)
    app.refresh_board(leaders_only=True)
    root.mainloop()


if __name__ == "__main__":
    main()


